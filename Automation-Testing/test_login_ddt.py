# Updated Signup DDT Automation
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, InvalidSessionIdException
from openpyxl import load_workbook
import time
import os

URL = "https://tichi-app-webapp-stage.web.app/"
EXCEL_FILE = r"D:\TichiAutomation\data\Automation_testing_ddt_data.xlsx"
SCREENSHOT_FOLDER = "screenshots"

os.makedirs(SCREENSHOT_FOLDER, exist_ok=True)

wb = load_workbook(EXCEL_FILE)
sheet = wb.active

if sheet["H1"].value is None:
    sheet["H1"] = "Actual Result"

if sheet["I1"].value is None:
    sheet["I1"] = "Status"

driver = webdriver.Chrome()
driver.maximize_window()
wait = WebDriverWait(driver, 15)

def take_screenshot(filename):
    try:
        driver.save_screenshot(os.path.join(SCREENSHOT_FOLDER, filename))
    except InvalidSessionIdException:
        pass
    except Exception:
        pass

for row in range(2, sheet.max_row + 1):

    first_name = str(sheet.cell(row,1).value or "")
    last_name = str(sheet.cell(row,2).value or "")
    phone = str(sheet.cell(row,3).value or "")
    email = str(sheet.cell(row,4).value or "")
    password = str(sheet.cell(row,5).value or "")
    confirm_password = str(sheet.cell(row,6).value or "")
    expected = str(sheet.cell(row,7).value or "").strip()

    actual = ""
    status = ""

    try:
        driver.get(URL)

        wait.until(
            EC.element_to_be_clickable(
                (By.XPATH,"//button[contains(text(),'Sign In')]")
            )
        ).click()

        email_box = wait.until(
            EC.visibility_of_element_located((By.ID,"email"))
        )
        email_box.clear()
        email_box.send_keys(email)

        driver.find_element(
            By.XPATH,
            "//button[@type='submit' and contains(.,'Continue')]"
        ).click()

        wait.until(
            EC.visibility_of_element_located((By.ID,"firstName"))
        )

        driver.find_element(By.ID,"firstName").send_keys(first_name)
        driver.find_element(By.ID,"lastName").send_keys(last_name)
        driver.find_element(By.ID,"phoneNumber").send_keys(phone)
        driver.find_element(By.ID,"password").send_keys(password)
        driver.find_element(By.ID,"confirmPassword").send_keys(confirm_password)
        driver.find_element(By.ID,"remember").click()

        driver.find_element(
            By.XPATH,
            "//button[@type='submit' and normalize-space()='Sign Up']"
        ).click()

        time.sleep(3)

        verify_heading = driver.find_elements(
            By.XPATH,
            "//h2[contains(text(),'Verify your account')]"
        )

        user_exists = driver.find_elements(
            By.XPATH,
            "//*[contains(text(),'User Already Exists')]"
        )

        if len(verify_heading) > 0:
            actual = "Success"
            status = "PASS"

        elif len(user_exists) > 0:
            actual = "Success"
            status = "PASS"

        else:
            error_message = ""

            errors = driver.find_elements(
                By.XPATH,
                "//p | //span | //small | //label | //div[contains(@class,'text-red')]"
            )

            for err in errors:
                txt = err.text.strip()
                if txt:
                    error_message = txt
                    break

            if error_message:
                actual = error_message
            else:
                actual = "Unknown Error"

            if actual.lower() == expected.lower():
                status = "PASS"
            else:
                status = "FAIL"
                take_screenshot(f"Row_{row}.png")

    except TimeoutException:
        actual = "Timeout"
        status = "PASS" if actual.lower() == expected.lower() else "FAIL"
        take_screenshot(f"Timeout_Row_{row}.png")

    except Exception as e:
        actual = str(e)
        status = "PASS" if actual.lower() == expected.lower() else "FAIL"
        take_screenshot(f"Exception_Row_{row}.png")

    sheet.cell(row,8).value = actual
    sheet.cell(row,9).value = status
    wb.save(EXCEL_FILE)

driver.quit()
wb.save(EXCEL_FILE)
print("Signup Automation Completed")
print("Excel Updated Successfully")

